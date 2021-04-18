#!/usr/bin/env python
# -*- coding: utf-8 -*-

"""
导出数据到excel
Created on 2021/04/17
@author: cjq
"""

from datetime import datetime

import openpyxl
import psycopg2
from openpyxl.styles import Font, numbers

from conf.database import batch_execute, execute_sql, get_data
from utils.time_utils import TimeDuration


def main(duration):
    """
    Exports data from PostgreSQL to an Excel spreadsheet using psycopg2.

    Arguments:
    headings - list of strings to use as column headings
    filepath - path and filename of the Excel file

    """

    headings = ('股票代码',	'证券简称',	'年度',	'所属行业',	'省',	'市',
                '销售收入',	'净利润',	'总资产',	'净资产',	'研发费用',	'政府补助')
    filepath = '.\export\企业研发补助{}.xlsx'.format(datetime.now().strftime("%Y%m%d%H%M%S"))
    print('{} 正在从数据库获取数据...'.format(datetime.now()))
    data = get_list()
    print('{} 成功获取数据，共计{}条，已耗时{}'.format(datetime.now(), len(data), duration.getTillNow()))
    wb = openpyxl.Workbook()
    sheet = wb.active

    sheet.row_dimensions[1].font = Font(bold=True)
    print('{} 准备导入到Excel...'.format(datetime.now()))
    # Spreadsheet row and column indexes start at 1
    # so we use "start = 1" in enumerate so
    # we don't need to add 1 to the indexes.
    for colno, heading in enumerate(headings, start=1):
        sheet.cell(row=1, column=colno).value = heading

    # This time we use "start = 2" to skip the heading row.
    for rowno, row in enumerate(data, start=2):
        for colno, cell_value in enumerate(row, start=1):
            cell = sheet.cell(row=rowno, column=colno)
            cell.value = cell_value
            if colno >= 6:
                cell.number_format = numbers.FORMAT_NUMBER_COMMA_SEPARATED1
    print('{} 导入Excel完成，准备保存文件到本地，已耗时{}'.format(datetime.now(), duration.getTillNow()))
    wb.save(filepath)
    print('{} 保存文件成功，已耗时{}'.format(datetime.now(), duration.getTillNow()))

def get_list():
    select_sql = 'SELECT t1.stock_code AS 股票代码, t1.bond_short AS 证券简称, t.year AS 年度, sw_industry_name AS 所属行业, t1.province AS 省, t1.city AS 市, amt1 AS 销售收入, amt2 AS 净利润, amt3 AS 总资产, amt4 AS 净资产, amt5 AS 研发费用, amt6 AS 政府补助 FROM t_jj_corp_research_subsidy t LEFT JOIN t_jj_corp_info t1 ON t1.id = t.corp_id WHERE 1=1 AND t.amt3 IS NOT NULL ORDER BY t1.stock_code, t.year DESC'
    return get_data(select_sql)


if __name__ == '__main__':
    duration = TimeDuration()
    duration.start()
    main(duration)
    duration.stop()
    duration.printDurationInfo()
